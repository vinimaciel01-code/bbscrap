from pathlib import Path

import os
import time
import pandas as pd
import numpy as np

from openpyxl import load_workbook

def salva_excel_nova_aba(filename, sheetname, dataframe, startrow=None):
    """
    Salva os dados em uma nova aba de um arquivo excel existente.
    Se já existir, cria nova aba com nome adaptado.
    O arquivo excel deve existir.
    """
    if startrow is None:
        startrow = 0
    if not isinstance(startrow, int):
        raise ValueError('Startrow deve ser "None" ou inteiro maior ou igual a zero.')
    if startrow < 0:
        raise ValueError('Startrow deve ser "None" ou inteiro maior ou igual a zero.')

    book = load_workbook(filename)
    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='w', startrow=None)
    writer.book = book
    dataframe.to_excel(writer, sheet_name=sheetname)
    writer.save()

def salva_excel_subst_aba(filename, sheetname, dataframe, startrow=None):
    """
    Salva os dados, substituindo uma aba existente. O arquivo excel deve existir.
    ExcelWriter usa writer.sheets para acessar a planilha. Se você deixar ele vazio,
    não saberá que a aba já existe e criará uma nova (adaptando o nome).
    @param filename: arquivo a salvar (path completo ou
    @param sheetname: nome da aba
    @param dataframe: dados a salvar
    @param startrow: linha superior a começar a salvar os dados
    """
    if startrow is None:
        startrow = 0
    if not isinstance(startrow, int):
        raise ValueError('Startrow deve ser "None" ou inteiro maior ou igual a zero.')
    if startrow < 0:
        raise ValueError('Startrow deve ser "None" ou inteiro maior ou igual a zero.')

    book = load_workbook(filename)
    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='w')
    writer.book = book

    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    dataframe.to_excel(writer, sheet_name=sheetname, engine='xlsxwriter', startrow=startrow)
    writer.save()

def salva_excel_novo_arquivo(filename, sheetname, dataframe, startrow=None):
    """
    Salva os dados na aba requerida, criando o arquivo (ou sobrescrevendo o existente)
    @param filename: arquivo a salvar (path completo ou
    @param sheetname: nome da aba
    @param dataframe: dados a salvar
    @param startrow: linha superior a começar a salvar os dados
    """
    if startrow is None:
        startrow = 0
    if not isinstance(startrow, int):
        raise ValueError('Startrow deve ser "None" ou inteiro maior ou igual a zero.')
    if startrow < 0:
        raise ValueError('Startrow deve ser "None" ou inteiro maior ou igual a zero.')

    writer = pd.ExcelWriter(filename, mode='w')
    dataframe.to_excel(writer, sheet_name=sheetname, engine='xlsxwriter', startrow=startrow)
    writer.save()

def salva_excel(filename, sheetname, dataframe, startrow=None):
    """
    Salva os dados na aba específica do arquivo excel. Caso já exista, sobrescreve. Caso o arquivo não exista, o cria.
    @param filename: arquivo a salvar (path completo ou relativo)
    @param sheetname: nome da aba
    @param dataframe: dados a salvar
    @param startrow:
        linha superior a começar a salvar os dados
    """
    if startrow is None:
        startrow = 0
    if not isinstance(startrow, int):
        raise ValueError('Startrow deve ser "None" ou inteiro maior ou igual a zero.')
    if startrow < 0:
        raise ValueError('Startrow deve ser "None" ou inteiro maior ou igual a zero.')

    if not os.path.isfile(filename):
        salva_excel_novo_arquivo(filename, sheetname, dataframe, startrow)

    salva_excel_subst_aba(filename, sheetname, dataframe, startrow)

def download_concluido(directory):
    dl_wait = True
    while dl_wait:
        time.sleep(1)
        dl_wait = False
        if is_download_finished(directory) is False:
            dl_wait = True

def is_download_finished(temp_folder):
    firefox_temp_file = sorted(Path(temp_folder).glob('*.part'))
    chrome_temp_file = sorted(Path(temp_folder).glob('*.crdownload'))
    downloaded_files = sorted(Path(temp_folder).glob('*.*'))
    if (len(firefox_temp_file) == 0) and \
            (len(chrome_temp_file) == 0) and \
            (len(downloaded_files) >= 1):
        return True
    else:
        return False


if __name__ == '__main__':

    x3 = np.random.randn(100, 2)
    df3 = pd.DataFrame(x3)

    x4 = np.random.randn(100, 2)
    df4 = pd.DataFrame(x4)

    salva_excel(r"../data/Consulta.xlsx", 'temp3', df4, startrow=None)
